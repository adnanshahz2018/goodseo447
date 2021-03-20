
#  Python imports
from os import replace
import threading
from selenium.webdriver.remote.webelement import WebElement
import xlsxwriter 
import json, time
import pandas as pd
from numpy import nan
import openpyxl as op
from bs4 import BeautifulSoup
from selenium import webdriver 


prefix = {'United States': 'https://www.amazon.com'}

class amazon:
    count = 0
    filename = 'product_data.xlsx'

    def __init__(self):
        self.create_excel_file()

    def scrape(self):
        browser = webdriver.Chrome('chromedriver.exe') 
        # url = 'https://www.amazon.com/s?i=specialty-aps&bbn=16225009011&rh=n%3A%2116225009011%2Cn%3A541966&ref=nav_em__nav_desktop_sa_intl_computers_and_accessories_0_2_5_6'
        url = 'https://www.amazon.com/s?i=specialty-aps&bbn=16225007011&rh=n%3A16225007011%2Cn%3A13896617011&ref=nav_em__nav_desktop_sa_intl_computers_tablets_0_2_6_4'
        browser.get(url)

        source = browser.page_source
        soup = BeautifulSoup(source, features='lxml')
        divs = soup.find_all('div', attrs={'class': 'a-section a-spacing-medium'})
        # our_div = divs[0]
        for our_div in divs:
            try:
                title = our_div.find('span', attrs={'class': 'a-size-base-plus a-color-base a-text-normal'}).get_text()
                div = our_div.find_all('div', attrs={'class':'a-section a-spacing-none a-spacing-top-small'})[0]
                h2 = div.find('h2')
                product_url = prefix['United States'] + h2.find('a')['href']
                self.get_info(browser, title, product_url)
            except: pass
            # time.sleep(3)
        browser.quit()

    def get_info(self, browser, title, product_url):
        browser.get(product_url)
        source = browser.page_source
        soup = BeautifulSoup(source, features='lxml')

        tr = soup.find('tr', attrs={'id': 'comparison_price_row'})
        td = tr.find('td', attrs={'class': 'comparison_baseitem_column'})
        span = td.find('span')
        price = float(str(span.find_all('span')[0].get_text()).replace('$',''))

        table = soup.find('table', attrs={'id':'productDetails_detailBullets_sections1'})
        tbody = table.find('tbody')
        tr = tbody.find_all('tr')[0]
        asin = str(tr.find('td').get_text()).replace('\n', '')

        print('\nTitle: ', title)
        print('Price: ', price)
        print('ASIN: ', asin)

        self.write_to_excel(title, product_url, price, asin)

    def create_excel_file(self):
        # creating the file for the first time 
        # if not os.path.exists(self.filename):
        workbook = xlsxwriter.Workbook(self.filename)
        worksheet = workbook.add_worksheet("data")
        workbook.close()
        wb = op.load_workbook(self.filename, False)
        ws = wb['data']
        ws.append(['Title','Product URL', 'Price', 'ASIN'])
        wb.save(self.filename)
        wb.close()

    def write_to_excel(self, title, product_url, price, asin):
        wb = op.load_workbook(self.filename, False)
        ws = wb['data']
        ws.append([title, product_url, price, asin])
        wb.save(self.filename)
        wb.close()


if __name__ == '__main__':
    amz = amazon()
    amz.scrape()